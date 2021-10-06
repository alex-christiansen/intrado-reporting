########################################################
### Import Packages
########################################################
import looker_sdk
from looker_sdk import models
import logging
import urllib3
import requests
import zipfile
import os
import csv
import glob   
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from PIL import ImageDraw


import re
import openpyxl
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl import Workbook
import csv
import io

from PIL import Image


# disable warnings coming from self-signed SSL cert
sdk = looker_sdk.init31(config_file='intradoint.ini') 
dashboard = sdk.dashboard(dashboard_id="2") 
filters = []
for filter in dashboard.dashboard_filters:
    if filter.default_value:
        filters.append(filter.name + ': ' + filter.default_value)

wb = Workbook()
ws = wb.active
ws.title = "Dashboard Details"
ft = Font(bold=True)

ws.cell(row = 1, column = 1, value = 'Business Pulse Summary').font = ft
ws['A3'] = 'Filters Used: '
ws['B3'] = '\n'.join(filters)
ws['A14'] = 'Generated at'
dateTimeObj = datetime.now()

ws['B14'] = str(dateTimeObj.year) + '/' +  str(dateTimeObj.month) +  '/' + str(dateTimeObj.day) + ' ' + str(dateTimeObj.hour) +  ':' + str(dateTimeObj.minute) + ':' + str(dateTimeObj.second)

ws.column_dimensions['A'].auto_size = True
ws.column_dimensions['B'].auto_size = True

img = openpyxl.drawing.image.Image('/usr/local/google/home/alchristiansen/python_scripts/dashboard-business_pulse/logo.png')
print('Image type of logo: ', type(img))
img.anchor = 'A2'
ws.add_image(img)
wb.save('./dashboard-business_pulse/pixel_perfect.xlsx')

dir = "./dashboard-business_pulse/pixel_perfect.xlsx"
book = load_workbook(dir)
writer = pd.ExcelWriter(dir, engine = 'openpyxl')
writer.book = book

dashboard_data = sdk.dashboard(str(2))
row = 0

# Insert an image.
for elements in dashboard_data.dashboard_elements:
    # if elements.result_maker and elements.result_maker.query_id == 18:
    if elements.result_maker and elements.result_maker.vis_config['type'] != 'single_value':
        df = io.StringIO(sdk.run_query(elements.result_maker.query_id, result_format='csv', apply_formatting=True))
        f = pd.read_csv(df,sep=',')
        f.to_excel(writer, sheet_name=elements.title[0:30], index = 0)
        workbook  = writer.book
        worksheet = writer.sheets[elements.title[0:30]]
        img_as_bytes = io.BytesIO(sdk.run_query(elements.result_maker.query_id, result_format='png'))
        img = openpyxl.drawing.image.Image(img_as_bytes)
        location = f.shape[0] + 5
        print(location)
        img.anchor = f"A{location}"
        worksheet.add_image(img)
    elif elements.result_maker:
        df = io.StringIO(sdk.run_query(elements.result_maker.query_id, result_format='csv', apply_formatting=True))
        f = pd.read_csv(df,sep=',')
        f.to_excel(writer, sheet_name='KPIs', index = 0, startrow = row)
        row = row + len(f.index) + 2

        # df = pd.read_csv(data, sep=',')
        # df.to_excel(writer, sheet_name='stuff', index=False)

        # print(elements.title + ': ' + str(elements.result_maker.query_id))
        # print(elements.result_maker.vis_config['type'])
# print(dashboard_data.dashboard_elements[0].result_maker.query_id)
# for elements in dashboard_data.dashboard_elements:
#     print(elements.result_maker)

writer.save()
writer.close()

wb = load_workbook(dir)

from openpyxl.utils import get_column_letter

# column_widths = []
# for row in data:
#     for i, cell in enumerate(row):
#         if len(column_widths) > i:
#             if len(cell) > column_widths[i]:
#                 column_widths[i] = len(cell)
#         else:
#             column_widths += [len(cell)]

ws = wb["KPIs"]
max_length = []
for row in ws.values:
    for value in row:
        if type(value) in [str]:
            max_length.append(len(value))  

ws.column_dimensions['A'].width = max(max_length)
ws.column_dimensions['B'].width = max(max_length)
ws.column_dimensions['C'].width = max(max_length)

      

wb.save(dir)



# for f in os.listdir(dir):
#     os.remove(os.path.join(dir, f))                                                                          

# with zipfile.ZipFile("../Downloads/dashboard-business_pulse.zip", 'r') as zip_ref:
#     zip_ref.extractall(".")

# all_files = glob.glob(os.path.join(dir, "*.csv"))
# for file in all_files:
#     new_name = os.path.split(file)[1].split('.')[0][0:31]





# for f in os.listdir(dir):
#     if f != 'out.xlsx':
#         os.remove(os.path.join(dir, f))     


