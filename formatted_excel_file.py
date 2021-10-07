########################################################
### Import Packages
########################################################
import looker_sdk
from looker_sdk import models
import logging, urllib3, requests, zipfile, os, csv, glob, re, io
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from PIL import ImageDraw, Image    

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Color, colors
from openpyxl.utils import get_column_letter


# disable warnings coming from self-signed SSL cert
class formatted_excel:
    def __init__(self, dashboard_id, logo_location, output_name):
        self.dashboard_id = dashboard_id
        self.logo_location = logo_location
        self.output_name = output_name
        self.sdk = looker_sdk.init31(config_file='intradoint.ini')
    
    def get_dashboard_date(self):
        dashboard = self.sdk.dashboard(self.dashboard_id)
        filters = []
        for filter in dashboard.dashboard_filters:
            if filter.default_value:
                filters.append(filter.name + ': ' + filter.default_value)
        return filters, dashboard

    def write_cover_page(self):
        print('Writing cover page...')
        dashboard_data = self.get_dashboard_date()
        wb = Workbook()
        ws = wb.active
        ws.title = 'Dashboard Details'
        ft = Font(bold=True)
        dateTimeObj = datetime.now()

        ws.cell(row=1, column=1, value=dashboard_data[1].title)
        ws['A3'] = 'Filters Used: '
        ws['B3'] = '\n'.join(dashboard_data[0])
        ws['A14'] = 'Generated at'
        ws['B14'] = str(dateTimeObj.year) + '/' +  str(dateTimeObj.month) +  '/' + str(dateTimeObj.day) + ' ' + str(dateTimeObj.hour) +  ':' + str(dateTimeObj.minute) + ':' + str(dateTimeObj.second)
        ws.column_dimensions['A'].auto_size = True
        ws.column_dimensions['B'].auto_size = True

        img = openpyxl.drawing.image.Image(self.logo_location)
        img.anchor = 'A2'
        ws.add_image(img)

        wb.save(self.output_name)
        wb.close()
        print('... cover page written')

    def add_sheets(self):
        dashboard_data = self.get_dashboard_date()
        book = load_workbook(self.output_name)
        writer = pd.ExcelWriter(self.output_name, engine = 'openpyxl')
        writer.book = book
        row = 0

        for elements in dashboard_data[1].dashboard_elements:
            if elements.result_maker and elements.result_maker.vis_config['type'] != 'single_value':
                print('Some table powered by query id: ', elements.result_maker.query_id)
                df = io.StringIO(self.sdk.run_query(elements.result_maker.query_id, result_format='csv', apply_formatting=True))
                f = pd.read_csv(df,sep=',')
                f.to_excel(writer, sheet_name=elements.title[0:30], index = 0)
                workbook  = writer.book
                worksheet = writer.sheets[elements.title[0:30]]
                img_as_bytes = io.BytesIO(self.sdk.run_query(elements.result_maker.query_id, result_format='png'))
                img = openpyxl.drawing.image.Image(img_as_bytes)
                location = f.shape[0] + 5
                print(location)
                img.anchor = f"A{location}"
                worksheet.add_image(img)
            elif elements.result_maker:
                print('Some single viz powered by query id: ', elements.result_maker.query_id)
                df = io.StringIO(self.sdk.run_query(elements.result_maker.query_id, result_format='csv', apply_formatting=True))
                f = pd.read_csv(df,sep=',')
                f.to_excel(writer, sheet_name='KPIs', index = 0, startrow = row)
                row = row + len(f.index) + 2
                
        writer.save()
        writer.close()

    def format_columns(self):
        wb = load_workbook(self.output_name)
        ws = wb["KPIs"]
        max_length = []
        for row in ws.values:
            for value in row:
                if type(value) in [str]:
                    max_length.append(len(value))  

        ws.column_dimensions['A'].width = max(max_length)
        ws.column_dimensions['B'].width = max(max_length)
        ws.column_dimensions['C'].width = max(max_length)
        wb.save(self.output_name)

    def main(self):
        self.write_cover_page()
        self.add_sheets()
        self.format_columns()

if __name__ == '__main__':
    db = formatted_excel(dashboard_id='2', logo_location='logo.png', output_name='./pixel_perfect_with_class.xlsx')
    db.main()
